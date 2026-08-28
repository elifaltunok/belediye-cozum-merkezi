import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from django.conf import settings


FONT_PATH = str(settings.BASE_DIR / 'static' / 'fonts' / 'DejaVuSans.ttf')
FONT_NAME = 'DejaVuSans'

try:
    pdfmetrics.registerFont(TTFont(FONT_NAME, FONT_PATH))
    TURKISH_FONT_AVAILABLE = True
except Exception:
    TURKISH_FONT_AVAILABLE = False
    FONT_NAME = 'Helvetica'


def build_tickets_excel(tickets):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Talepler"

    headers = [
        'Takip Kodu', 'Başlık', 'Kategori', 'İlçe', 'Mahalle', 'Durum',
        'Birim', 'Telefon', 'E-posta', 'Destek Sayısı', 'Oluşturulma', 'Güncellenme',
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for t in tickets:
        ws.append([
            t.tracking_code, t.title, t.category.name, t.district, t.neighborhood or '',
            t.get_status_display(), t.get_current_unit_display() if t.current_unit else '',
            t.phone or '', t.email or '', t.support_count,
            t.created_at.strftime('%d.%m.%Y %H:%M'), t.updated_at.strftime('%d.%m.%Y %H:%M'),
        ])

    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(header) + 4, 14)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_dashboard_pdf(context):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleTR', parent=styles['Title'], fontName=FONT_NAME, fontSize=18)
    heading_style = ParagraphStyle('HeadingTR', parent=styles['Heading2'], fontName=FONT_NAME, fontSize=13, spaceBefore=14, spaceAfter=6)
    normal_style = ParagraphStyle('NormalTR', parent=styles['Normal'], fontName=FONT_NAME, fontSize=10)

    elements = []
    elements.append(Paragraph("Akıllı Kent Sahada — Yönetici Özet Raporu", title_style))
    elements.append(Paragraph(f"Oluşturulma Tarihi: {context['generated_at']}", normal_style))
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("Genel Özet", heading_style))
    summary_data = [
        ['Toplam Talep', str(context['total_tickets'])],
        ['Beklemede', str(context['pending_count'])],
        ['Çözülen', str(context['resolved_count'])],
        ['Ortalama Çözüm Süresi (saat)', str(context['avg_resolution_hours'] or '—')],
    ]
    summary_table = Table(summary_data, colWidths=[8*cm, 6*cm])
    summary_table.setStyle(_table_style())
    elements.append(summary_table)

    elements.append(Paragraph("Durum Dağılımı", heading_style))
    status_data = [['Durum', 'Adet']] + context['status_rows']
    status_table = Table(status_data, colWidths=[8*cm, 6*cm])
    status_table.setStyle(_table_style(header=True))
    elements.append(status_table)

    elements.append(Paragraph("İlçe Bazlı Talep Sayısı", heading_style))
    district_data = [['İlçe', 'Adet']] + context['district_rows']
    district_table = Table(district_data, colWidths=[8*cm, 6*cm])
    district_table.setStyle(_table_style(header=True))
    elements.append(district_table)

    elements.append(Paragraph("Birim Performansı", heading_style))
    unit_data = [['Birim', 'Çözüm Kaydı Sayısı']] + context['unit_rows']
    unit_table = Table(unit_data, colWidths=[8*cm, 6*cm])
    unit_table.setStyle(_table_style(header=True))
    elements.append(unit_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def build_ticket_pdf(ticket, resolutions):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleTR', parent=styles['Title'], fontName=FONT_NAME, fontSize=16)
    heading_style = ParagraphStyle('HeadingTR', parent=styles['Heading2'], fontName=FONT_NAME, fontSize=12, spaceBefore=12, spaceAfter=6)
    normal_style = ParagraphStyle('NormalTR', parent=styles['Normal'], fontName=FONT_NAME, fontSize=10, leading=14)

    elements = []
    elements.append(Paragraph(f"Talep Kaydı — {ticket.tracking_code}", title_style))
    elements.append(Spacer(1, 10))

    info_data = [
        ['Başlık', ticket.title],
        ['Kategori', ticket.category.name],
        ['İlçe / Mahalle', f"{ticket.district} / {ticket.neighborhood or '—'}"],
        ['Durum', ticket.get_status_display()],
        ['Birim', ticket.get_current_unit_display() if ticket.current_unit else '—'],
        ['Oluşturulma', ticket.created_at.strftime('%d.%m.%Y %H:%M')],
    ]
    info_table = Table(info_data, colWidths=[4*cm, 12*cm])
    info_table.setStyle(_table_style())
    elements.append(info_table)

    elements.append(Paragraph("Açıklama", heading_style))
    elements.append(Paragraph(ticket.description, normal_style))

    if resolutions:
        elements.append(Paragraph("İşlem Geçmişi", heading_style))
        for r in resolutions:
            text = f"<b>{r.created_at.strftime('%d.%m.%Y %H:%M')}</b> — {r.get_new_status_display()}"
            if r.handled_by:
                text += f" ({r.handled_by.username})"
            elements.append(Paragraph(text, normal_style))
            if r.note:
                elements.append(Paragraph(r.note, normal_style))
            elements.append(Spacer(1, 6))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def _table_style(header=False):
    style = [
        ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]
    if header:
        style.append(('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')))
        style.append(('TEXTCOLOR', (0, 0), (-1, 0), colors.white))
        style.append(('FONTNAME', (0, 0), (-1, 0), FONT_NAME))
    else:
        style.append(('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f4f6fb')))
    return TableStyle(style)